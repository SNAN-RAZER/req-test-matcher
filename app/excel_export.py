from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.models import AnalysisReport

FILLS = {
    "aligned": "C6EFCE",
    "partial": "FFEB9C",
    "wrong": "FFC7CE",
    "orphan": "BDD7EE",
    "uncovered": "E2D5F1",
    "header": "1F2A36",
}


def report_to_xlsx(report: AnalysisReport) -> bytes:
    wb = Workbook()
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )
    head_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    def sheet(title: str, headers: list[str], rows: list[list[object]], verdict_col: int | None = None):
        ws = wb.create_sheet(title)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col, h)
            cell.font = head_font
            cell.fill = PatternFill("solid", fgColor=FILLS["header"])
        for r_i, row in enumerate(rows, 2):
            for c_i, val in enumerate(row, 1):
                cell = ws.cell(r_i, c_i, val)
                cell.alignment = wrap
                cell.border = thin
                if verdict_col is not None and c_i == verdict_col:
                    color = FILLS.get(str(val).lower())
                    if color:
                        cell.fill = PatternFill("solid", fgColor=color)
        widths = [18] * len(headers)
        for row in rows:
            for i, val in enumerate(row):
                widths[i] = min(60, max(widths[i], min(60, len(str(val)) + 2)))
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        return ws

    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Requirement ↔ test knowledge graph"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A3"] = report.summary
    summary["A5"] = "Chat model"
    summary["B5"] = report.model
    summary["A6"] = "Vision model"
    summary["B6"] = report.vision_model
    summary["A7"] = "Requirements"
    summary["B7"] = len(report.requirements)
    summary["A8"] = "Test cases"
    summary["B8"] = len(report.test_cases)
    summary["A9"] = "Wrong tests"
    summary["B9"] = len(report.wrong_tests)
    summary["A10"] = "Orphan tests"
    summary["B10"] = len(report.orphan_tests)
    summary["A11"] = "Uncovered requirements"
    summary["B11"] = len(report.uncovered_requirements)
    summary["A12"] = "Graph nodes"
    summary["B12"] = report.knowledge_graph.node_count
    summary["A13"] = "Graph edges"
    summary["B13"] = report.knowledge_graph.edge_count
    summary["A15"] = "Files processed"
    for i, name in enumerate(report.files_processed, 16):
        summary.cell(i, 1, name)
    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 80

    sheet(
        "Traceability",
        ["Test ID", "Requirement ID", "Verdict", "Graph score", "Rationale", "Issues"],
        [
            [
                m.test_id,
                m.requirement_id or "",
                m.verdict,
                round(m.score, 3),
                m.rationale,
                "; ".join(m.issues),
            ]
            for m in report.matches
        ],
        verdict_col=3,
    )
    sheet(
        "Requirements",
        ["ID", "Title", "Text", "Source", "Parent"],
        [[r.id, r.title, r.text, r.source_file, r.parent_file] for r in report.requirements],
    )
    sheet(
        "Test cases",
        ["ID", "Title", "Text", "Source", "Linked IDs"],
        [
            [
                t.id,
                t.title,
                t.text,
                t.source_file,
                ", ".join(t.extra.get("linked_requirement_ids") or [])
                if isinstance(t.extra.get("linked_requirement_ids"), list)
                else str(t.extra.get("linked_requirement_ids") or ""),
            ]
            for t in report.test_cases
        ],
    )
    sheet(
        "Wrong tests",
        ["Test ID", "Requirement ID", "Rationale", "Issues"],
        [[m.test_id, m.requirement_id or "", m.rationale, "; ".join(m.issues)] for m in report.wrong_tests],
    )
    sheet(
        "Uncovered",
        ["Requirement ID"],
        [[rid] for rid in report.uncovered_requirements],
    )
    sheet(
        "Knowledge graph nodes",
        ["ID", "Kind", "Label", "Title", "Verdict"],
        [
            [n.get("id"), n.get("kind"), n.get("label"), n.get("title"), n.get("verdict")]
            for n in report.knowledge_graph.nodes
        ],
    )
    sheet(
        "Knowledge graph edges",
        ["Source", "Target", "Relation"],
        [
            [e.get("source"), e.get("target"), e.get("rel")]
            for e in report.knowledge_graph.edges
        ],
    )
    sheet(
        "Agent swarm",
        ["Agent", "Role", "Action", "Detail"],
        [[t.agent, t.role, t.action, t.detail] for t in report.agent_trace],
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
